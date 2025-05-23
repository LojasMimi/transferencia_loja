import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
import os

# ================
# CONFIGURAÇÃO INICIAL
# ================
st.set_page_config(page_title="Transferência entre Lojas", layout="wide")
st.title("📦 Transferência entre Lojas")

# ================
# ESTADO DA SESSÃO
# ================
if "formulario_dados" not in st.session_state:
    st.session_state.formulario_dados = []

if "codigos_lote_processados" not in st.session_state:
    st.session_state.codigos_lote_processados = False

# ================
# SELEÇÃO DE LOJAS
# ================
st.header("🔄 Seleção de Lojas")

col1, col2 = st.columns([1, 1])
with col1:
    lojas = ["MIMI", "KAMI", "TOTAL MIX", "E-COMMERCE"]
    de_loja = st.selectbox("De qual loja está saindo a transferência?", lojas)
with col2:
    para_lojas = [loja for loja in lojas if loja != de_loja]
    para_loja = st.selectbox("Para qual loja vai a transferência?", para_lojas)

# ================
# LEITURA DE PLANILHAS DO REPOSITÓRIO
# ================
st.header("1️⃣ Planilhas de Produtos Carregadas")

df_list = []
planilhas_path = "planilhas"  # pasta com os arquivos no repo

if not os.path.exists(planilhas_path):
    st.error(f"⚠️ Pasta '{planilhas_path}' não encontrada no repositório.")
    st.stop()

arquivos_planilhas = [f for f in os.listdir(planilhas_path) if f.endswith(('.xlsx', '.xls'))]

if not arquivos_planilhas:
    st.warning("⚠️ Não foram encontradas planilhas na pasta 'planilhas'. Por favor, faça upload pelo formulário.")
    st.stop()

for arquivo in arquivos_planilhas:
    caminho_arquivo = os.path.join(planilhas_path, arquivo)
    try:
        excel = pd.ExcelFile(caminho_arquivo)
        all_sheets_df = []
        for sheet_name in excel.sheet_names:
            df = pd.read_excel(caminho_arquivo, sheet_name=sheet_name, header=13, usecols="A:P")
            df.columns = [col.strip().upper() for col in df.columns]
            df['__ORIGEM_PLANILHA__'] = f"{arquivo} - {sheet_name}"
            all_sheets_df.append(df)
        combined_df = pd.concat(all_sheets_df, ignore_index=True)
        df_list.append(combined_df)
        st.success(f"✅ Planilha '{arquivo}' processada com sucesso.")
    except Exception as e:
        st.error(f"Erro ao processar a planilha '{arquivo}': {e}")
        df_list.append(pd.DataFrame())

planilhas_validas = [df for df in df_list if not df.empty]
if not planilhas_validas:
    st.warning("⚠️ Nenhuma planilha válida carregada.")
    st.stop()

df_combinado_total = pd.concat(planilhas_validas, ignore_index=True)

with st.expander("👀 Visualização da Planilha Combinada"):
    st.dataframe(df_combinado_total, use_container_width=True)

# ================
# SELEÇÃO DE MODO
# ================
st.header("2️⃣ Escolha a Ação Desejada")
acao = st.radio("O que você deseja fazer?", ["FAZER TRANSFERENCIA", "RECEBER/CONFERIR TRANSFERENCIA"], horizontal=True)

# ================
# FAZER TRANSFERÊNCIA
# ================
if acao == "FAZER TRANSFERENCIA":
    modo = st.radio("Como deseja preencher o formulário?", ["Preenchimento Individual", "Preenchimento em Lote"], horizontal=True)

    if modo == "Preenchimento em Lote":
        with st.expander("📥 Upload da Planilha com os Códigos de Barras"):
            # Botão de download do modelo - colocado em coluna para responsividade
            download_col1, _ = st.columns([1, 4])
            with download_col1:
                if st.button("📥 Baixar Formulário de Lote"):
                    modelo_df = pd.DataFrame(columns=["CODIGO BARRA", "QUANTIDADE"])
                    buffer_modelo = BytesIO()
                    with pd.ExcelWriter(buffer_modelo, engine='openpyxl') as writer:
                        modelo_df.to_excel(writer, index=False, sheet_name="FormularioLote")
                    buffer_modelo.seek(0)
                    st.download_button("⬇️ Clique aqui para baixar o modelo de formulário", buffer_modelo, "modelo_formulario_lote.xlsx")

            codigos_file = st.file_uploader("📑 Envie a planilha preenchida", type=["xlsx", "xls"], key="codigos_lote")
            if codigos_file and not st.session_state.codigos_lote_processados:
                try:
                    codigos_df = pd.read_excel(codigos_file)
                    codigos_df.columns = [col.strip().upper() for col in codigos_df.columns]
                    if "CODIGO BARRA" not in codigos_df.columns or "QUANTIDADE" not in codigos_df.columns:
                        st.error("A planilha deve conter as colunas 'CODIGO BARRA' e 'QUANTIDADE'.")
                    else:
                        encontrados, nao_encontrados = [], []
                        for _, row in codigos_df.iterrows():
                            codigo = str(row["CODIGO BARRA"])
                            quantidade = int(row["QUANTIDADE"]) if pd.notna(row["QUANTIDADE"]) else 1
                            encontrado = False
                            for df in df_list:
                                resultado = df[df["CODIGO BARRA"].astype(str) == codigo]
                                if not resultado.empty:
                                    r = resultado.iloc[0]
                                    encontrados.append({
                                        "CODIGO BARRA": codigo,
                                        "CODIGO": r.get("CODIGO", ""),
                                        "FORNECEDOR": r.get("FORNECEDOR", ""),
                                        "DESCRICAO": r.get("DESCRIÇÃO", ""),
                                        "QUANTIDADE": quantidade
                                    })
                                    encontrado = True
                                    break
                            if not encontrado:
                                nao_encontrados.append({"CODIGO BARRA": codigo})

                        st.session_state.formulario_dados.extend(encontrados)
                        st.session_state.codigos_lote_processados = True

                        if encontrados:
                            st.success("✅ Códigos encontrados:")
                            st.dataframe(pd.DataFrame(encontrados), use_container_width=True)
                        if nao_encontrados:
                            st.warning("⚠️ Códigos não encontrados:")
                            st.dataframe(pd.DataFrame(nao_encontrados), use_container_width=True)

                except Exception as e:
                    st.error(f"Erro: {e}")

    else:  # Preenchimento Individual
        with st.expander("🔍 Buscar Produto e Adicionar"):
            tipo, valor, qtd = st.columns([2, 4, 2])
            with tipo:
                busca_tipo = st.selectbox("Buscar por:", ["Código de Barras", "Código do Fornecedor (REF)"], key="busca_tipo")
            with valor:
                busca_valor = st.text_input("Digite o valor:", key="busca_valor")
            with qtd:
                busca_qtd = st.number_input("Quantidade:", min_value=1, step=1, key="busca_qtd")

            if st.button("🔎 Procurar"):
                encontrado = False
                for df in df_list:
                    col_busca = "CODIGO BARRA" if busca_tipo == "Código de Barras" else "CODIGO"
                    resultado = df[df[col_busca].astype(str) == busca_valor.strip()]
                    if not resultado.empty:
                        r = resultado.iloc[0]
                        if len(st.session_state.formulario_dados) < 30:
                            st.session_state.formulario_dados.append({
                                "CODIGO BARRA": str(r.get("CODIGO BARRA", "")),
                                "CODIGO": r.get("CODIGO", ""),
                                "FORNECEDOR": r.get("FORNECEDOR", ""),
                                "DESCRICAO": r.get("DESCRIÇÃO", ""),
                                "QUANTIDADE": busca_qtd
                            })
                            st.success("✅ Produto adicionado!")
                            st.write(r)
                        else:
                            st.error("🚫 Máximo de 30 itens atingido.")
                        encontrado = True
                        break
                if not encontrado:
                    st.warning("Produto não encontrado.")

    # Exibir formulário e gerar relatório
    if st.session_state.formulario_dados:
        with st.expander(f"📋 Itens no Formulário ({len(st.session_state.formulario_dados)} itens)"):
            st.dataframe(pd.DataFrame(st.session_state.formulario_dados), use_container_width=True)

        if st.button("📄 Gerar Relatório Excel"):
            try:
                wb = load_workbook("FORMULÁRIO DE TRANSFERENCIA ENTRE LOJAS.xlsx")
                ws = wb.active

                # Preenchendo cabeçalho
                ws["A4"] = f"DE: {de_loja}"
                ws["C4"] = para_loja
                ws["D4"] = "DATA " + datetime.today().strftime("%d/%m/%Y")

                # Preenchendo itens
                for i, item in enumerate(st.session_state.formulario_dados[:30]):
                    ws[f"A{8+i}"].value = item["CODIGO BARRA"]
                    ws[f"A{8+i}"].number_format = '@'
                    ws[f"B{8+i}"] = item["CODIGO"]
                    ws[f"C{8+i}"] = item["FORNECEDOR"]
                    ws[f"D{8+i}"] = item["DESCRICAO"]
                    ws[f"E{8+i}"] = item["QUANTIDADE"]

                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                st.download_button("⬇️ Baixar Formulário Preenchido", buffer, "FORMULARIO_PREENCHIDO.xlsx")
            except Exception as e:
                st.error(f"Erro ao gerar o relatório: {e}")

# ================
# RECEBER/CONFERIR TRANSFERÊNCIA
# ================
else:
    with st.expander("📤 Upload do Formulário de Transferência para Conferência"):
        formulario_file = st.file_uploader("Envie o formulário preenchido", type=["xlsx", "xls"], key="formulario_upload")

        if formulario_file:
            try:
                formulario_df = pd.read_excel(formulario_file, header=6)
                formulario_df.columns = [col.strip().upper() for col in formulario_df.columns]

                st.subheader("📋 Dados do Formulário")
                st.dataframe(formulario_df, use_container_width=True)

                if "CODIGO DE BARRAS" not in formulario_df.columns:
                    st.error("❌ A coluna 'CODIGO DE BARRAS' não foi encontrada.")
                    st.stop()

                codigos = formulario_df["CODIGO DE BARRAS"].dropna().astype(str).str.strip()
                df_filtrado = df_combinado_total[df_combinado_total["CODIGO BARRA"].astype(str).str.strip().isin(codigos)]

                if df_filtrado.empty:
                    st.warning("⚠️ Nenhum produto do formulário encontrado nas planilhas.")
                else:
                    st.success(f"✅ {len(df_filtrado)} produto(s) encontrados:")
                    st.dataframe(df_filtrado, use_container_width=True)
            except Exception as e:
                st.error(f"Erro ao processar o formulário: {e}")

# ================
# RODAPÉ / COPYRIGHT
# ================
st.markdown(
    """
    <hr style='border: 0; height: 1px; background: #ccc; margin-top: 2em; margin-bottom: 1em;' />
    <div style='text-align: center; color: grey; font-size: 0.8em;'>
        Aplicativo desenvolvido por <strong>PABLO</strong> para as lojas <strong>MIMI</strong>. Todos os direitos reservados.
    </div>
    """,
    unsafe_allow_html=True
)
